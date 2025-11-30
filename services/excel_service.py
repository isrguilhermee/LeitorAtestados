from openpyxl import Workbook, load_workbook
import os


class ExcelService:
    """Service for managing Excel files to store medical certificate data."""
    
    def __init__(self, filename="atestados.xlsx"):
        """
        Initialize Excel service.
        
        Args:
            filename: Name of the Excel file (default: 'atestados.xlsx')
        """
        package_root = os.path.dirname(os.path.dirname(__file__))
        self.filename = os.path.join(package_root, filename)
        self.column_headers = [
            "CID",
            "Médico",
            "Data de Emissão",
            "Dias de Repouso",
        ]
    
    def save_data(self, data):
        """
        Save medical certificate data to Excel file.
        
        Args:
            data: Dictionary with medical certificate information
                 Expected keys: CID, Médico, Data de Início, Dias de Afastamento
        """
        if os.path.exists(self.filename):
            wb = load_workbook(self.filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(self.column_headers)

        # Ensure header reflects the latest structure
        if ws.max_row >= 1:
            existing_headers = [cell.value for cell in ws[1]]
            if existing_headers != self.column_headers:
                for idx, header in enumerate(self.column_headers, start=1):
                    ws.cell(row=1, column=idx, value=header)
        
        ws.append([
            data.get("CID", ""),
            data.get("Médico", ""),
            data.get("Data de Emissão", ""),
            data.get("Dias de Repouso", ""),
        ])
        
        wb.save(self.filename)
    
    def get_all_data(self):
        """
        Read all data from the Excel file.
        
        Returns:
            list: List of dictionaries containing all records
        """
        if not os.path.exists(self.filename):
            return []
        
        wb = load_workbook(self.filename)
        ws = wb.active
        
        data = []
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell for cell in row):  # Skip empty rows
                values = list(row)
                data.append({
                    "CID": values[0] if len(values) > 0 else "",
                    "Médico": values[1] if len(values) > 1 else "",
                    "Data de Emissão": values[2] if len(values) > 2 else "",
                    "Dias de Repouso": values[3] if len(values) > 3 else "",
                })
        
        return data
    
    def create_file(self):
        """
        Create a new Excel file with headers.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(self.column_headers)
            wb.save(self.filename)
            return True
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            return False
    
    def file_exists(self):
        """
        Check if the Excel file exists.
        
        Returns:
            bool: True if file exists, False otherwise
        """
        return os.path.exists(self.filename)


# Legacy function for backward compatibility
def save_to_excel(data, filename="atestados.xlsx"):
    """Legacy function wrapper for backward compatibility."""
    excel = ExcelService(filename)
    excel.save_data(data)
